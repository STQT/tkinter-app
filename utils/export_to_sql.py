import os
import sqlite3
import openpyxl

def export_to_sqlite():
	# 1. Создание и подключение к базе
	
	# Получаем текущую папку проекта
	prj_dir = os.path.abspath(os.path.curdir)
	
	a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
	
	# Имя базы
	base_name = 'sql_export.db'
	
	# метод sqlite3.connect автоматически создаст базу, если ее нет
	connect = sqlite3.connect(prj_dir + '/' + base_name)
	# курсор - это специальный объект, который делает запросы и получает результаты запросов
	cursor = connect.cursor()
	
	# создание таблицы если ее не существует
	cursor.execute('CREATE TABLE IF NOT EXISTS cars (brand text, model text, distance int , year int)')
	
	# 2. Работа c xlsx файлом
	
	# Читаем файл и лист1 книги excel
	file_to_read = openpyxl.load_workbook('Cars.xlsx', data_only=True)
	sheet = file_to_read['Sheet1']
	
	# Цикл по строкам начиная со второй (в первой заголовки)
	
	for row in range(2, sheet.max_row + 1):
		# Объявление списка
		data = []
		# Цикл по столбцам от 1 до 4 ( 5 не включая)
		for col in range(1, 5):
			# value содержит значение ячейки с координатами row col
			value = sheet.cell(row, col).value
			# Список который мы потом будем добавлять
			data.append(value)
		
		# 3. Запись в базу и закрытие соединения
		
		# Вставка данных в поля таблицы
		cursor.execute("INSERT INTO cars VALUES (?, ?, ?, ?);", (data[i] for i in range(1,5)))
	
	# сохраняем изменения
	connect.commit()
	# закрытие соединения
	connect.close()
	
def clear_base():
	'''Очистка базы sqlite'''
	
	# получаем текущую папку проекта
	prj_dir = os.path.abspath(os.path.curdir)
	
	# Имя базы
	base_name = 'sql_export.db'
	
	connect = sqlite3.connect(prj_dir + '/' + base_name)
	cursor = connect.cursor()
	
	# Запись в базу, сохранение и закрытие соединения
	cursor.execute("DELETE FROM cars")
	connect.commit()
	connect.close()
	
export_to_sqlite()
